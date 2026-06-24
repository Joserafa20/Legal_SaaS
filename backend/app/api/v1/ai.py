import os, uuid, json
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel
from typing import Optional, List
from app.core.database import get_db
from app.core.deps import get_current_user, require_role
from app.services.ai_service import ai_service
from app.models.user import User

router = APIRouter(prefix="/ai", tags=["Asistente IA"])


class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    message: str
    messages: Optional[List[ChatMessage]] = None
    context: Optional[dict] = None


class DocumentAnalysisRequest(BaseModel):
    case_context: Optional[str] = ""
    document_text: Optional[str] = ""


class LegalResearchRequest(BaseModel):
    query: str
    jurisdiction: Optional[str] = "CGP"


class DraftRequest(BaseModel):
    type: str
    caseId: Optional[str] = ""
    instructions: Optional[str] = ""


@router.post("/extract")
async def extract_file_text(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    content = await file.read()
    ext = os.path.splitext(file.filename or "")[1].lower()
    text = ""

    try:
        if ext == ".pdf":
            import fitz
            doc = fitz.open(stream=content, filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
            if len(text.strip()) < 50:
                import io, easyocr, numpy as np
                reader = easyocr.Reader(['es'], gpu=False, verbose=False)
                doc2 = fitz.open(stream=content, filetype="pdf")
                texts = []
                for page in doc2:
                    pix = page.get_pixmap(dpi=200)
                    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
                    result = reader.readtext(img, detail=0, paragraph=True)
                    texts.append("\n".join(result))
                doc2.close()
                text = "\n\n".join(texts)
        elif ext == ".docx":
            import io
            from docx import Document
            doc = Document(io.BytesIO(content))
            text = "\n".join(p.text for p in doc.paragraphs)
        elif ext == ".xlsx":
            import io
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            rows = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    rows.append(" | ".join(str(c) if c is not None else "" for c in row))
            text = "\n".join(rows)
            wb.close()
        else:
            text = content.decode("utf-8", errors="replace")
    except Exception as e:
        text = f"[Error al extraer texto: {str(e)}]"

    text = text[:10000]
    return {"filename": file.filename, "text": text, "file_type": file.content_type}


@router.post("/chat")
async def chat(
    request: ChatRequest,
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    try:
        if not ai_service.api_key or ai_service.api_key == "sk-or-v1-placeholder":
            return {
                "data": {
                    "id": "",
                    "role": "assistant",
                    "content": "El asistente IA no está configurado. Configure la variable OPENROUTER_API_KEY en el archivo .env del backend para habilitar esta funcionalidad.",
                    "timestamp": ""
                }
            }
        system_msg = {"role": "system", "content": "Eres un asistente juridico experto en legislacion colombiana. Responde preguntas legales con fundamento. Cita articulos de ley y jurisprudencia cuando sea posible. Responde en espanol. Escribe en texto simple. NO uses asteriscos, numerales, guiones para formato, ni ningun simbolo de Markdown. Para enumerar usa numeros seguidos de punto (1. 2. 3.)."}
        history = [{"role": m.role, "content": m.content} for m in (request.messages or []) if m.role != "system"]
        messages = [system_msg] + history + [{"role": "user", "content": request.message}]
        response = await ai_service.chat_completion(messages)
        return {
            "data": {
                "id": "",
                "role": "assistant",
                "content": response,
                "timestamp": ""
            }
        }
    except Exception as e:
        return {
            "data": {
                "id": "",
                "role": "assistant",
                "content": f"Error al procesar la consulta: {str(e)}",
                "timestamp": ""
            }
        }


@router.get("/history")
async def get_chat_history(
    caseId: Optional[str] = Query(None),
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    return {"data": []}


@router.delete("/history")
async def clear_chat_history(
    caseId: Optional[str] = Query(None),
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    return {"data": None}


@router.post("/analyze/{document_id}")
async def analyze_document(
    document_id: str,
    request: DocumentAnalysisRequest = DocumentAnalysisRequest(),
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    try:
        text = request.document_text or f"Documento {document_id}"
        messages = [
            {
                "role": "system",
                "content": "Eres un analista jurídico. Analiza el documento proporcionado y extrae los puntos clave."
            },
            {
                "role": "user",
                "content": text
            }
        ]
        response = await ai_service.chat_completion(messages)
        return {
            "data": {
                "summary": response,
                "keyPoints": []
            }
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error: {str(e)}"
        )


@router.post("/research")
async def legal_research(
    request: LegalResearchRequest,
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    try:
        messages = [
            {
                "role": "system",
                "content": f"Eres un investigador jurídico experto en legislación colombiana ({request.jurisdiction}). Proporciona resultados de investigación detallados."
            },
            {
                "role": "user",
                "content": request.query
            }
        ]
        response = await ai_service.chat_completion(messages)
        return {
            "data": {
                "results": [response],
                "sources": []
            }
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error: {str(e)}"
        )


@router.post("/draft")
async def draft_document(
    request: DraftRequest,
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    try:
        messages = [
            {
                "role": "system",
                "content": "Eres un abogado experto en redacción de documentos legales colombianos."
            },
            {
                "role": "user",
                "content": f"Tipo: {request.type}. Instrucciones: {request.instructions}. Caso ID: {request.caseId}"
            }
        ]
        response = await ai_service.chat_completion(messages)
        return {
            "data": {
                "content": response
            }
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error: {str(e)}"
        )


class CaseAnalysisRequest(BaseModel):
    case_context: str
    question: str
    jurisdiction: str = "CGP"


class MemorialDraftRequest(BaseModel):
    case_context: str
    memorial_type: str
    jurisdiction: str = "CGP"
    instructions: Optional[str] = ""


class JurisprudenceRequest(BaseModel):
    legal_topic: str
    case_context: str


class ViabilityAnalysisRequest(BaseModel):
    case_context: str
    proposed_action: str
    jurisdiction: str = "CGP"


@router.post("/analyze-case")
async def analyze_case(
    request: CaseAnalysisRequest,
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    try:
        response = await ai_service.analyze_case(
            case_context=request.case_context,
            question=request.question,
            jurisdiction=request.jurisdiction
        )
        return {"analysis": response}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al analizar caso: {str(e)}"
        )


@router.post("/draft-memorial")
async def draft_memorial(
    request: MemorialDraftRequest,
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    try:
        response = await ai_service.draft_memorial(
            case_context=request.case_context,
            memorial_type=request.memorial_type,
            jurisdiction=request.jurisdiction,
            instructions=request.instructions
        )
        return {"draft": response}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar borrador: {str(e)}"
        )


@router.post("/suggest-jurisprudence")
async def suggest_jurisprudence(
    request: JurisprudenceRequest,
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    try:
        response = await ai_service.suggest_jurisprudence(
            legal_topic=request.legal_topic,
            case_context=request.case_context
        )
        return {"jurisprudence": response}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al buscar jurisprudencia: {str(e)}"
        )


@router.post("/analyze-viability")
async def analyze_viability(
    request: ViabilityAnalysisRequest,
    current_user: User = Depends(require_role(["admin", "abogado"]))
):
    try:
        response = await ai_service.analyze_viability(
            case_context=request.case_context,
            proposed_action=request.proposed_action,
            jurisdiction=request.jurisdiction
        )
        return {"viability": response}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al analizar viabilidad: {str(e)}"
        )
